from . import *
import os
import json
import redis

from incoming.models import Accessory
from lhwms.settings import BASE_DIR
from django.http import FileResponse

import openpyxl
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, colors, Alignment
# json.dumps(list(data), cls=DjangoJSONEncoder),序列化数据类型操作

'''
===================================
公共函数: 附件上传、excel下载
===================================
'''


def attachment_uploading(request, model):
    """前端传递："pk"-->queryset "id"；"accessory"：上传的文件列表
    (附件上传操作，文件保存到本地MEDIA_ROOT)
    :param model: Accessory外键链接模型对象
    :return:-->dict:文件上传结果信息
    """
    pk = request.POST.get('pk')  # 获取文件存储连接对象id值
    incoming = model.objects.get(pk=pk)  # 获取被上传文件对象
    accessorys = request.FILES.getlist('accessory')  # 获取前端传来的文件对象
    data = {}  # 建立文件上传状态信息dict
    for f in accessorys:
        # 判断表中"accessory"是否存在上传的文件,没有保存到数据库中，django自动保存到MEDIA_ROOT
        obj, create = Accessory.objects.get_or_create(accessory=f)
        # (< Accessory: Accessory object >, False)
        if create:  # 上传操作了
            obj.table_id = incoming  # 文件外键连接到当前incoming对象
            obj.save()
            data.update({str(f): "上传成功!"})
        else:
            data.update({str(f): "该文件已存在，请删除后再上传!"})

    return data


def attachment_delete(request):
    """
    附件删除操作
    :param request:
    :return: data
    """
    pk = request.POST.get('pk')
    files_obj = Accessory.objects.filter(table_id=pk)
    files_name = files_obj.values('accessory')
    data = {}
    for f in files_name:  # 删除dir文件
        DIR = os.path.join(BASE_DIR, 'media', str(f['accessory']))
        # print(DIR)
        if os.path.exists(DIR):
            os.remove(DIR)
            data.update({str(f['accessory']): '该文件已删除'})
        else:
            data.update({str(f['accessory']): '该文件不存在'})
    for f in files_obj:  # 删除数据库dir
        f.delete()
    return data


def file_iterator(file_name, chunk_size=512):
    """文件读取生成器"""
    with open(file_name, 'rb') as f:
        while True:
            c = f.read(chunk_size)
            if c:
                yield c
            else:
                break


def attachment_download(request):
    """
    下载附件，pk是一个，可以多个迭代下载
    :param request:
    :return: attachment
    """
    pk = request.POST.get('pk')
    file_name = str(Accessory.objects.get(pk=pk).accessory)
    # 问题.txt <class 'django.db.models.fields.files.FieldFile'>
    file_dir = os.path.join(BASE_DIR, 'media', file_name)

    file = open(file_dir, 'rb')

    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename={}'. \
        format(file_name).encode('utf-8')

    return response


def export_excel(request, model, query_mark, filename,
                 sheet_name, sheet_head_data):
    """前端传递值：None;
    (下载Excel表格);
    :param model:-->model: 被导出excel模型;
    :param query_mark:-->str: 模块名称;
    :param filename: 客户端下载excel文件命名。（本接口使用query_mark固定值）;
    :param sheet_name: excel "sheet"命名;
    :param sheet_head_data:-->dict: "sheet"第一行数据，需要字段名，列宽，颜色等信息（可选）
    :return: return FileResponse(filename) or "msg" if rc."key" is None;
    """
    excel_abs_dir = os.path.join(BASE_DIR, 'media', 'excels', filename)
    max_row_num = 100000  # 最大导出行数
    # 从缓存加载数据
    rc = redis.Redis(connection_pool=pool_result)
    key = get_key(request, model, query_mark)  # 生成组合"key"值
    if rc.get(key) is None:
        msg = '未查询到导出数据，请查询后重试！'
        return msg
    else:
        data_rc = json.loads(rc.get(key))  # 获取redis缓存中数据

        # excel填充数据
        data = [[x for x in sheet_head_data.keys()]]  # 列表解析加载"sheet"标题行
        # queryset数据转化data = [['1', '邓平', '25'], ['2', '云云', '23']]
        for sheet_num, data_dic in enumerate(data_rc):
            data_lst = [sheet_num+1]  # 序号num
            for value in data_dic.values():
                data_lst.append(value)
            data.append(data_lst)
        # 限制最大导出行数
        if len(data) > max_row_num:
            data = data[: max_row_num]

        workbook = openpyxl.Workbook()  # 创建文件对象
        create_sheet = workbook.create_sheet(sheet_name, 0)  # 初始位置创建sheet对象
        # 设置表头背景色
        fp = PatternFill(fill_type='solid', fgColor="77DDFF")  # 浅蓝色

        for index, width in enumerate([x for x in sheet_head_data.values()]):
            create_sheet.cell(row=1, column=index+1).fill = fp  # 第一行加浅蓝色
            cm = get_column_letter(index+1)  # 根据序列生成对应列标号，第一列
            create_sheet.column_dimensions[cm].width = width  # 对应字段名设置列宽

        for row_index, row in enumerate(data):  # 数据写入到单元格中
            for col_index, col in enumerate(row):
                # 对"sheet"cell数据写入
                create_sheet.cell(row_index+1, col_index+1).value = col
                # 字体设置，颜色等，样式
                # sheet.cell(row_index+1, col_index+1).font = Font(
                #     color=colors.BLUE
                # )
                # 设置文字位置，对其方式=左对齐
                create_sheet.cell(row_index+1, col_index+1).alignment = Alignment(
                    horizontal='left',  # 水平：center left，right
                    vertical='center'  # 垂直: top,bottom,center
                )

        del workbook['Sheet']  # 删除workbook中初始'Sheet'表名
        workbook.save(excel_abs_dir)  # 保存文件到本地, save调用了close

        # 返回文件流于浏览器
        file = open(excel_abs_dir, 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename={}'. \
            format(filename).encode('utf-8')  # filename可以另外写一个函数动态生成
        return response
